import os
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# --- 1. Generate Invoice PDF ---
class InvoicePDF(FPDF):
    def __init__(self):
        super().__init__()
        font_path = None
        candidates = [
            "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
            "/System/Library/Fonts/Hiragino Sans GB.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        ]
        for p in candidates:
            if os.path.exists(p):
                font_path = p
                break

        if font_path:
            self.add_font("JP", "", font_path, uni=True)
            self.add_font("JP", "B", font_path, uni=True)
            self.jp_font = "JP"
        else:
            self.add_font("Helvetica", "", "")
            self.jp_font = "Helvetica"

def generate_invoice_pdf():
    pdf = InvoicePDF()
    pdf.add_page()
    f = pdf.jp_font

    pdf.set_font(f, "B", 20)
    pdf.cell(0, 15, "請 求 書", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font(f, "", 10)
    pdf.cell(0, 6, "請求書番号: INV-2025-00042", ln=True, align="R")
    pdf.cell(0, 6, "請求日: 2025-04-15", ln=True, align="R")
    pdf.cell(0, 6, "支払期日: 2025-05-31", ln=True, align="R")
    pdf.ln(5)

    pdf.set_font(f, "B", 12)
    pdf.cell(0, 8, "請求先", ln=True)
    pdf.set_font(f, "", 10)
    pdf.cell(0, 6, "株式会社スノーフレーク・ジャパン", ln=True)
    pdf.cell(0, 6, "東京都港区六本木6-10-1 六本木ヒルズ森タワー", ln=True)
    pdf.ln(5)

    pdf.set_font(f, "B", 12)
    pdf.cell(0, 8, "請求元", ln=True)
    pdf.set_font(f, "", 10)
    pdf.cell(0, 6, "合同会社テックソリューションズ", ln=True)
    pdf.cell(0, 6, "東京都渋谷区神宮前1-2-3", ln=True)
    pdf.cell(0, 6, "TEL: 03-1234-5678", ln=True)
    pdf.cell(0, 6, "担当: 山田太郎", ln=True)
    pdf.ln(8)

    pdf.set_font(f, "B", 12)
    pdf.cell(0, 8, "明細", ln=True)

    items = [
        ("クラウド基盤構築サービス", 1, 1500000),
        ("データ移行作業", 3, 300000),
        ("セキュリティ監査", 1, 500000),
        ("運用サポート（月額）", 2, 200000),
        ("トレーニング研修", 5, 80000),
    ]

    col_widths = [80, 25, 35, 35]
    headers = ["品名", "数量", "単価", "金額"]

    pdf.set_font(f, "B", 10)
    pdf.set_fill_color(220, 230, 241)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font(f, "", 10)
    subtotal = 0
    for name, qty, price in items:
        amount = qty * price
        subtotal += amount
        pdf.cell(col_widths[0], 8, name, border=1)
        pdf.cell(col_widths[1], 8, str(qty), border=1, align="C")
        pdf.cell(col_widths[2], 8, f"{price:,}", border=1, align="R")
        pdf.cell(col_widths[3], 8, f"{amount:,}", border=1, align="R")
        pdf.ln()

    tax = int(subtotal * 0.10)
    total = subtotal + tax

    pdf.ln(3)
    pdf.set_font(f, "", 10)
    pdf.cell(105, 7, "")
    pdf.cell(35, 7, "小計", border=1, align="C")
    pdf.cell(35, 7, f"{subtotal:,} 円", border=1, align="R")
    pdf.ln()
    pdf.cell(105, 7, "")
    pdf.cell(35, 7, "消費税(10%)", border=1, align="C")
    pdf.cell(35, 7, f"{tax:,} 円", border=1, align="R")
    pdf.ln()
    pdf.set_font(f, "B", 11)
    pdf.cell(105, 8, "")
    pdf.cell(35, 8, "合計", border=1, align="C", fill=True)
    pdf.cell(35, 8, f"{total:,} 円", border=1, align="R", fill=True)
    pdf.ln(15)

    pdf.set_font(f, "", 9)
    pdf.cell(0, 6, "備考: お振込手数料はご負担ください。", ln=True)
    pdf.cell(0, 6, "振込先: みずほ銀行 渋谷支店 普通 1234567 ゴウドウガイシャテックソリューションズ", ln=True)

    out_path = os.path.join(OUTPUT_DIR, "invoice_sample.pdf")
    pdf.output(out_path)
    print(f"Generated: {out_path}")
    return out_path


# --- 2. Generate Excel file ---
def generate_excel():
    wb = Workbook()

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    # Sheet 1: Monthly Sales
    ws1 = wb.active
    ws1.title = "月次売上"

    headers1 = ["月", "売上（万円）", "コスト（万円）", "利益（万円）", "利益率(%)"]
    data1 = [
        ("2025年1月", 4500, 3200, 1300, 28.9),
        ("2025年2月", 5200, 3500, 1700, 32.7),
        ("2025年3月", 6800, 4100, 2700, 39.7),
        ("2025年4月", 5900, 3800, 2100, 35.6),
        ("2025年5月", 7200, 4300, 2900, 40.3),
        ("2025年6月", 8100, 4800, 3300, 40.7),
    ]

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(data1, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 2:
                cell.alignment = Alignment(horizontal="right")

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    # Sheet 2: Department breakdown
    ws2 = wb.create_sheet("部門別売上")
    headers2 = ["部門", "Q1売上（万円）", "Q2売上（万円）", "年間目標（万円）", "達成率(%)"]
    data2 = [
        ("営業部", 12000, 15800, 50000, 55.6),
        ("マーケティング部", 3500, 4200, 15000, 51.3),
        ("エンジニアリング部", 8000, 9500, 30000, 58.3),
        ("カスタマーサクセス部", 2800, 3100, 10000, 59.0),
        ("データサイエンス部", 1200, 1800, 8000, 37.5),
    ]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 2:
                cell.alignment = Alignment(horizontal="right")

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20

    out_path = os.path.join(OUTPUT_DIR, "sales_data.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")
    return out_path


if __name__ == "__main__":
    generate_invoice_pdf()
    generate_excel()
    print("\nDone! Sample files generated.")
